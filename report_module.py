import os
from datetime import datetime
from collections import Counter
import difflib
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import matplotlib.pyplot as plt

def get_output(sheet, column, exclude_strings):
    return sum(1 for row in range(1, sheet.max_row + 1)
               if (val := sheet.cell(row=row, column=column).value) and val not in exclude_strings)

def get_accepted_rings(sheet, column_letter, string_to_find):
    return sum(1 for row in range(1, sheet.max_row + 1) if sheet[f"{column_letter}{row}"].value == string_to_find)

def get_rejected_rings(sheet, column, exclude_strings):
    return sum(1 for row in range(1, sheet.max_row + 1)
               if (val := sheet.cell(row=row, column=column).value) and val not in exclude_strings)

def get_rework_rings(sheet, column_letter, string_to_find):
    return sum(1 for row in range(1, sheet.max_row + 1) if sheet[f"{column_letter}{row}"].value == string_to_find)

def calculate_yield(okay_rings, total_rings):
    return (okay_rings / total_rings) * 100 if total_rings else 0

def get_rejection_details(sheet, column_index):
    values = [cell.value for row in sheet.iter_rows(min_col=column_index, max_col=column_index, min_row=1, max_row=sheet.max_row)
              for cell in row if cell.value and str(cell.value).strip()]
    return dict(Counter(values))

def get_cover_mismatch(sheet, column_letter, string_to_find):
    return sum(1 for row in range(1, sheet.max_row + 1) if sheet[f"{column_letter}{row}"].value == string_to_find)

def fuzzy_match(reason, keywords, cutoff=0.8):
    match = difflib.get_close_matches(reason, keywords, n=1, cutoff=cutoff)
    return match[0] if match else None

def generate_bar_chart(data_dict, output_folder):
    plt.figure(figsize=(12, 6))
    labels, values = zip(*data_dict.items()) if data_dict else ([], [])
    bars = plt.bar(labels, values, color="#3498db")

    for bar in bars:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2.0, yval + 0.5, f"{int(yval)}", ha='center', va='bottom', fontsize=9)

    plt.xticks(rotation=45, ha='right')
    plt.ylabel("Count")
    plt.title("Individual Rejection Reasons")
    plt.tight_layout()

    chart_path = os.path.join(output_folder, "rejection_chart.png")
    plt.savefig(chart_path)
    plt.close()
    return chart_path

def generate_report(file_path, report_for):
    today = datetime.today().date()
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    total_rings = get_output(sheet, 3, [])
    accepted_rings = get_accepted_rings(sheet, 'C', 'Accepted')
    rejected_rings = get_rejected_rings(sheet, 3, ['Accepted', 'REWORK', 'Cover Mismatch'])
    reworked_rings = get_rework_rings(sheet, 'C', 'REWORK')
    yield_percentage = calculate_yield(accepted_rings, total_rings)
    rejection_details = get_rejection_details(sheet, 4)
    cover_mismatch = get_cover_mismatch(sheet, 'D', 'Cover Mismatch')

    for key in ['Accepted', 'REWORK', 'Cover Mismatch']:
        rejection_details.pop(key, None)

    casting_keywords = ["MICRO BUBBLES","COIL MISALIGNMENT","DENT ON RESIN","DUST INSIDE RESIN","RESIN CURING ISSUE",
                        "SHORT FILL OF RESIN","SPM REJECTION","TIGHT FIT FOR CHARGE"]

    assembly_keywords = ["BLACK GLUE", "ULTRAHUMAN TEXT SMUDGED","WHITE PATCH ON INSERT","WHITE PATCH ON PCB",
                         "WHITE PATCH ON TAPE NEAR BATTERY","WRONG RX COIL"]

    polishing_keywords = ['SHELL COATING REMOVED', 'SIDE SCRATCH', 'SIDE SCRATCH(EMERY)',
                          'IMPROPER RESIN FINISH', 'RESIN DAMAGE','LOOSE FITTING ON CHARGER',
                          'RX COIL SCRACTH', 'SCRATCHES ON RESIN', 'UNEVEN POLISHING',
                          "SCRATCHES ON SHELL & SIDE SHELL"]

    shell_keywords = ["BLACK MARKS ON SHELL","DENT ON SHELL","DISCOLORATION","IRREGULAR SHELL SHAPE",
                      "SHELL COATING ISSUE","WHITE MARKS ON SHELL"]

    functional_keywords = ['100% ISSUE','3 SENSOR ISSUE','BATTERY ISSUE','BLUETOOTH HEIGHT ISSUE','CE TAPE ISSUE',
                           'CHARGING CODE ISSUE','COIL THICKNESS ISSUE/BATTERY THICKNESS','COMPONENT HEIGHT ISSUE',
                           'CURRENT ISSUE','DISCONNECTING ISSUE','HRS BUBBLE','HRS COATING HEIGHT ISSUE',
                           'HRS DOUBLE LIGHT ISSUE','NO NOTIFICATION IN CDT','NOT ADVERTISING (WINGLESS PCB)',
                           'NOT CHARGING','SENSOR ISSUE','STC ISSUE']

    def lower_all(lst): return [kw.lower() for kw in lst]
    casting_keywords = lower_all(casting_keywords)
    assembly_keywords = lower_all(assembly_keywords)
    polishing_keywords = lower_all(polishing_keywords)
    shell_keywords = lower_all(shell_keywords)
    functional_keywords = lower_all(functional_keywords)

    assembly_rejections, casting_rejections = {}, {}
    polishing_rejections, shell_rejections = {}, {}
    functional_rejections, other_rejections = {}, {}

    for reason, count in rejection_details.items():
        reason_clean = str(reason).strip()
        reason_lower = reason_clean.lower()
        if fuzzy_match(reason_lower, assembly_keywords):
            assembly_rejections[reason_clean] = count
        elif fuzzy_match(reason_lower, casting_keywords):
            casting_rejections[reason_clean] = count
        elif fuzzy_match(reason_lower, polishing_keywords):
            polishing_rejections[reason_clean] = count
        elif fuzzy_match(reason_lower, shell_keywords):
            shell_rejections[reason_clean] = count
        elif fuzzy_match(reason_lower, functional_keywords):
            functional_rejections[reason_clean] = count
        else:
            other_rejections[reason_clean] = count

    # Write report to XLSX
    wb_out = Workbook()
    ws = wb_out.active
    bold = Font(bold=True)
    row = 1
    ws[f"A{row}"] = f"REPORT FOR {report_for.upper()}: {today}"
    ws[f"A{row}"].font = bold
    row += 2

    def write_row(label, value):
        nonlocal row
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = bold
        ws[f"B{row}"] = value
        row += 1

    write_row("OUTPUT", total_rings)
    write_row("OKAY", accepted_rings)
    write_row("REJECTED", rejected_rings)
    if reworked_rings > 0:
        write_row("REWORK", reworked_rings)
    write_row("YIELD", f"{yield_percentage:.2f}%")
    row += 1

    def write_rejections(title, data_dict):
        nonlocal row
        ws[f"A{row}"] = title
        ws[f"A{row}"].font = bold
        row += 1
        if data_dict:
            for reason, count in data_dict.items():
                ws[f"A{row}"] = reason
                ws[f"B{row}"] = count
                row += 1
        else:
            ws[f"A{row}"] = "None"
            row += 1
        row += 1

    write_rejections("ASSEMBLY REJECTIONS", assembly_rejections)
    write_rejections("CASTING REJECTIONS", casting_rejections)
    write_rejections("FUNCTIONAL REJECTIONS", functional_rejections)
    write_rejections("POLISHING REJECTIONS", polishing_rejections)
    write_rejections("SHELL REJECTIONS", shell_rejections)
    write_rejections("OTHER REJECTIONS", other_rejections)

    ws[f"A{row}"] = "COVER MISMATCH"
    ws[f"A{row}"].font = bold
    ws[f"B{row}"] = cover_mismatch

    output_folder = os.path.dirname(file_path)
    output_path = os.path.join(output_folder, "output.xlsx")
    wb_out.save(output_path)

    chart_path = generate_bar_chart(rejection_details, output_folder)
    os.startfile(output_path)
    os.startfile(chart_path)
    return output_path
