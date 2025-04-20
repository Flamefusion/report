import pandas as pd
from collections import Counter
import openpyxl
from datetime import datetime

excel_file_path = r'c:\Users\meshe\OneDrive\Desktop\SB.xlsx'

# Get today's date
today = datetime.today().date()

# A function to count all the rows in the table - to get the output(total rings)
def get_output(sheet):
    return sheet.max_row

# A function to count all the accepted rings - to get the amount of okay rings
def get_accepted_rings(sheet, column_letter, string_to_find):
    count = 0
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet[f"{column_letter}{row}"].value
        if cell_value == string_to_find:
            count += 1
    return count

# A function to count all the rejection rings - to get the amount of rejected rings
def get_rejected_rings(sheet, column, exclude_strings):
    count = 0
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column).value
        if cell_value not in exclude_strings:
            count += 1
    return count

# A function to count all the rework rings - to get the amount of rework rings
def get_rework_rings(sheet, column_letter, string_to_find):
    count = 0
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet[f"{column_letter}{row}"].value
        if cell_value == string_to_find:
            count += 1
    return count

# A function to calculate the overall yield - to get the overall yield = [okay rings / total rings] * 100
def calculate_yield(okay_rings, total_rings):
    return (okay_rings / total_rings) * 100

# A function to fetch all the rejection details from the sheet and store in a dictionary
def get_rejection_details(sheet, column_index):
    column_values = []
    
    # Iterate through all rows in the specified column by index
    for row in sheet.iter_rows(min_col=column_index, max_col=column_index, min_row=1, max_row=sheet.max_row):
        for cell in row:
            column_values.append(cell.value)
    
    # Count the occurrences of each value using Counter
    value_counts = Counter(column_values)
    
    # Filter out values that appear more than once (repeated values)
    repeated_values = {value: count for value, count in value_counts.items() if count > 1}
    
    return repeated_values

def main():
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active  # Use the active sheet of the workbook
    
    total_rings = get_output(sheet)
    accepted_rings = get_accepted_rings(sheet, 'C', 'ACCEPTED') 
    rejected_rings = get_rejected_rings(sheet, 3, ['ACCEPTED', 'REWORK'])
    reworked_rings = get_rework_rings(sheet, 'C', 'REWORK')
    yield_percentage = calculate_yield(accepted_rings, total_rings)
    rejection_details = get_rejection_details(sheet, 3)

    # Exclude accepted and rework from the rejection details dictionary
    rejection_details.pop('ACCEPTED', None)  # Use pop to avoid KeyError
    rejection_details.pop('REWORK', None)    # Use pop to avoid KeyError
    

    # Convert the rejection details to a DataFrame
    df = pd.DataFrame.from_dict(rejection_details, orient='index')
    
    # Write the report to a text file insted of shoing in the console
    with open("C:/Users/meshe/OneDrive/Desktop/report/output.txt", "w") as f:

        print("REPORT FOR 3DE TECH: ",today,file=f)
        print(f"OUTPUT: {total_rings}",file=f)
        print(f"OKAY: {accepted_rings}",file=f)
        print(f"REJECTED: {rejected_rings}",file=f)
        if reworked_rings > 0:
            print(f"REWORK: {reworked_rings}",file=f)
        else:
            pass
        print(f"YIELD: {yield_percentage:.2f}%",file=f)
        print(f"YIELD: {yield_percentage:.2f}%",file=f)

        print("\nREJECTION DETAILS:",file=f)
        print("Casting Rejections -",file=f)

        # Print the DataFrame without the column header (by setting header=False when printing)
        print(df.to_string(header=False),file=f)

'''
spearate the generated dictionary to 3 different sub dictionaries to sort 3 different tupe of rejections ie : casting, shell and others.
then we can use dictionary indexing to get specific values we need or print full dictionary in data frame. can do this only when i have a proper data sheet.
'''

if __name__ == '__main__':
    main()
